from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from dift.reports.models import DiffReport


def render_excel(report: DiffReport, output: str | None = None) -> Path:
    """
    Render and write an Excel report.

    Excel reports should always be written to a file.
    """

    output_path = Path(output or "dift_report.xlsx")
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"

    _add_summary_sheet(ws, report)
    _add_quality_sheet(wb, report)
    _add_numeric_sheet(wb, report)
    _add_outlier_sheet(wb, report)
    _add_categorical_sheet(wb, report)

    wb.save(output_path)
    return output_path


def _add_summary_sheet(ws, report: DiffReport) -> None:
    rows = [
        ["Metric", "Value"],
        ["old_rows", report.summary.old_rows],
        ["new_rows", report.summary.new_rows],
        ["row_delta", report.summary.row_delta],
        ["risk_level", report.summary.risk_level],
    ]

    for row in rows:
        ws.append(row)

    _style_header(ws)
    _auto_size_columns(ws)
    ws.freeze_panes = "A2"


def _add_quality_sheet(wb: Workbook, report: DiffReport) -> None:
    ws = wb.create_sheet("Quality Diff")

    rows = [
        [
            "Column",
            "Old Nulls",
            "New Nulls",
            "Old Null %",
            "New Null %",
            "Delta Null %",
            "Spike",
            "Severity",
        ]
    ]

    for item in report.quality_diff.null_diffs:
        rows.append(
            [
                item.column,
                item.old_nulls,
                item.new_nulls,
                item.old_null_pct,
                item.new_null_pct,
                item.delta_null_pct,
                "Yes" if item.is_spike else "No",
                item.severity,
            ]
        )

    duplicate = report.quality_diff.duplicate_diff

    rows.extend(
        [
            [],
            ["Duplicate Metric", "Value"],
            ["Old duplicates", duplicate.old_duplicates],
            ["New duplicates", duplicate.new_duplicates],
            ["Delta duplicates", duplicate.delta_duplicates],
            ["Old duplicate %", duplicate.old_duplicate_pct],
            ["New duplicate %", duplicate.new_duplicate_pct],
            ["Delta duplicate %", duplicate.delta_duplicate_pct],
            ["Duplicate basis", duplicate.duplicate_basis],
            ["Spike", "Yes" if duplicate.is_spike else "No"],
            ["Severity", duplicate.severity],
        ]
    )

    for row in rows:
        ws.append(row)

    _style_header(ws)
    _auto_size_columns(ws)
    ws.freeze_panes = "A2"


def _add_numeric_sheet(wb: Workbook, report: DiffReport) -> None:
    ws = wb.create_sheet("Numeric Drift")

    rows = [
        [
            "Column",
            "Old Mean",
            "New Mean",
            "Delta Mean",
            "Mean Shift %",
            "Old Std",
            "New Std",
            "Delta Std",
            "Std Shift %",
            "Delta Range",
            "Range Shift %",
            "Threshold",
            "Drifted",
            "Severity",
        ]
    ]

    for item in report.numeric_diff:
        rows.append(
            [
                item.column,
                item.old_mean,
                item.new_mean,
                item.delta_mean,
                item.mean_shift_pct,
                item.old_std,
                item.new_std,
                item.delta_std,
                item.std_shift_pct,
                item.delta_range,
                item.range_shift_pct,
                item.drift_threshold,
                "Yes" if item.is_drifted else "No",
                item.severity,
            ]
        )

    for row in rows:
        ws.append(row)

    _style_header(ws)
    _auto_size_columns(ws)
    ws.freeze_panes = "A2"


def _add_outlier_sheet(wb: Workbook, report: DiffReport) -> None:
    ws = wb.create_sheet("Outlier Diff")

    rows = [
        [
            "Column",
            "Method",
            "Old Outliers",
            "New Outliers",
            "Delta Outliers",
            "Old Outlier %",
            "New Outlier %",
            "Delta Outlier %",
            "Lower Bound",
            "Upper Bound",
            "Spike",
            "Severity",
        ]
    ]

    for item in report.outlier_diff:
        rows.append(
            [
                item.column,
                item.method,
                item.old_outliers,
                item.new_outliers,
                item.delta_outliers,
                item.old_outlier_pct,
                item.new_outlier_pct,
                item.delta_outlier_pct,
                item.lower_bound,
                item.upper_bound,
                "Yes" if item.is_spike else "No",
                item.severity,
            ]
        )

    for row in rows:
        ws.append(row)

    _style_header(ws)
    _auto_size_columns(ws)
    ws.freeze_panes = "A2"


def _add_categorical_sheet(wb: Workbook, report: DiffReport) -> None:
    ws = wb.create_sheet("Categorical Diff")

    rows = [
        [
            "Column",
            "Values Added",
            "Values Removed",
            "Frequency Shifts",
            "Max Frequency Shift",
            "Shifted",
            "Severity",
        ]
    ]

    for item in report.categorical_diff:
        rows.append(
            [
                item.column,
                ", ".join(item.values_added),
                ", ".join(item.values_removed),
                _format_frequency_shifts(item.frequency_shifts),
                item.max_frequency_shift,
                "Yes" if item.is_shifted else "No",
                item.severity,
            ]
        )

    for row in rows:
        ws.append(row)

    _style_header(ws)
    _auto_size_columns(ws)
    ws.freeze_panes = "A2"


def _format_frequency_shifts(shifts: dict[str, float]) -> str:
    if not shifts:
        return ""

    return ", ".join(f"{value}: {shift:.2%}" for value, shift in shifts.items())


def _style_header(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F2937")
    header_font = Font(color="FFFFFF", bold=True)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font


def _auto_size_columns(ws) -> None:
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)

        for cell in column_cells:
            value = "" if cell.value is None else str(cell.value)
            max_length = max(max_length, len(value))

        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
